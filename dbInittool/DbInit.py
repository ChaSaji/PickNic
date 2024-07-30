from openpyxl import load_workbook
import json

namelist = {
    "Meal",
    "Badge",
    "MealStatus",
    "RecipeDetail",
    "Material",
    "MaterialPhotoRelation",
    "Photo",
    "User"
}


def read_excel_to_list(filename):
    filename = 'excels/'+filename + '.xlsx'
    print(filename)
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    data = []
    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, cell_value in zip(headers, row):
            row_data[header] = cell_value
        data.append(row_data)

    return data


def save_to_file(data, excelname, output_file):
    with open(output_file, 'a') as file:
        file.write("export const " + excelname + " =")
        json.dump(data, file, indent=4)
        file.write(";"+'\n')


#output_file = 'dataBaseInit.js'  # 出力ファイルのパスを指定
output_file = '../nativeApp/src/lib/dataBaseInit.js'  # 出力ファイルのパスを指定
#output_file ="output.js"
with open(output_file, 'w') as file:
    file.write("//reset\n")

for excelname in namelist:
    result = read_excel_to_list(excelname)
    print(result)
    save_to_file(result, excelname, output_file)
